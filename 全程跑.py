








import requests
import openpyxl

def read_data(filename,sheetname):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    max_row = sheet.max_row
    case_list=[]
    for i in range(2,max_row+1):
        dict1 = dict(
        case_id= sheet.cell(row=i,column=1).value,
        url = sheet.cell(row=i,column=5).value,
        data = sheet.cell(row=i,column=6).value,
        expect = sheet.cell(row=i,column=7).value,
        )
        case_list.append(dict1)
    return case_list

def login_QCD(url,data):
  headers_login = {'X-Lemonban-Media-Type': 'lemonban.v2', 'Content-Type': 'application/json'}
  login = requests.post(url=url, headers=headers_login, json=data)
  re =  login.json()
  return re

def write_result(filename,sheetname,row,column,final_result):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row, column=column).value = final_result
    wb.save(filename)


cases = read_data('test_case_api.xlsx','register')          #调用读取测试用例，获取所有测试用例数据保存到变量

def result(filename,sheetname):
    cases = read_data(filename,sheetname)
    for case in cases:
        case_id = case.get('case_id')
        url = case.get('url')
        data = eval(case.get('data'))                       #eval运行被字符串包裹的表达式，可以去掉字符串
        expect = eval(case.get('expect'))                     #获取预期结果
        expect_msg = expect.get('msg')                        #获取预期结果的msg
        # print(type(data))                                      #判断类型
        real_result = login_QCD(url=url,data=data)           #调用发送接口请求函数，返回结果用变量real_result接受
        real_msg = real_result.get('msg')                    #获取实际的msg
        print('预期结果的msg：{}'.format(expect_msg))
        print('实际结果的msg:{}'.format(real_msg))
        if real_msg == expect_msg:
            print('{},pass'.format(case_id))
            final_re = 'pass'
        else:
            print('{},false'.format(case_id))
            final_re = 'false'
        write_result(filename,sheetname,case_id+1,8,final_re)

result('test_case_api.xlsx','login')